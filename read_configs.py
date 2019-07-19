#!/usr/bin/env python3
# -*- coding:utf-8 -*-
# @Time : 2019/1/7 下午7:05 
# @Author : zhubc 
# @File : read_configs.py 
# @Software: PyCharm
import csv
from xml.etree import ElementTree as ET
import openpyxl


def read_config(filepath, nodepath):
    '''when config xml contains tag and text,return a dict like {"tag": "text"}'''
    try:
        et = ET.parse(filepath)
    except FileNotFoundError:
        raise Exception("配置文件不存在:" + filepath)
    else:
        datas = et.findall(nodepath)
        if len(datas) > 0:
            data_dic = {}
            for data in datas:
                data_dic[data.tag] = data.text
            return data_dic
        else:
            raise Exception("未找到配置信息")


def read_case_config(filepath, nodepath):
    '''when config xml only contains tag,return a list like [tag1, tag2]'''
    try:
        et = ET.parse(filepath)
    except FileNotFoundError:
        raise Exception("配置文件不存在:" + filepath)
    else:
        datas = et.findall(nodepath)
        if len(datas) > 0:
            data_list = []
            for data in datas:
                data_list.append(data.tag)
            return data_list
        else:
            raise Exception("未找到测试用例配置信息")


def read_excel_config(filepath, sheetname):
    try:
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.get_sheet_by_name(sheetname)
    except Exception:
        raise Exception("EXCEL配置信息不存在:" + filepath + ' ' + sheetname)
    else:
        nrows = sheet.max_row
        if nrows > 0:
            li = []
            for n in range(2, nrows+1):
                row_datas = sheet[n]
                row_data = {}
                for i in range(1, len(row_datas+1)):
                    row_data[sheet.cell(1, i).value] = sheet.cell(n, i).value
                else:
                    li.append(row_data)

            return {sheetname: li}
        else:
            raise Exception("配置文件为空")


def read_csv_config(filepath):
    try:
        with open(filepath, 'r', encoding='utf-8-sig', errors='ignore') as f:
            csv_reader = csv.reader(f)
            data = []
            next(csv_reader)
            for i in csv_reader:
                data.append(i)
            return data
    except Exception:
        raise Exception("用例文件不存在：" + filepath)


if __name__ == '__main__':
    from pprint import pprint
    # pprint(read_excel_config('./excel_cases/case.xlsx', '用例'))
    print(read_csv_config('./excel_cases/case.csv'))